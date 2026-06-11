"""
Importa datos de inventario desde un archivo Excel (.xlsx).
Uso: python scripts/import_excel.py <archivo.xlsx> <barrio_id>

El Excel debe tener al menos las columnas: nombre, categoria, ubicacion
Columnas opcionales: descripcion, estado, cantidad, marca, modelo, numero_serie, notas
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from app import create_app, db
from app.models import Item, Categoria, Barrio, Historial, User

app = create_app()


def importar(archivo, barrio_id, user_id=1):
    barrio = db.session.get(Barrio, barrio_id)
    if not barrio:
        print(f"ERROR: Barrio con id={barrio_id} no encontrado.")
        return

    wb = load_workbook(archivo)
    ws = wb.active

    # Leer headers de la primera fila
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    def col_idx(name):
        """Busca la columna por nombre (case insensitive, parcial)."""
        for i, h in enumerate(headers):
            if name in h:
                return i
        return None

    # Mapeo de columnas
    col_nombre = col_idx("nombre")
    col_codigo = col_idx("codigo")
    col_cat = col_idx("categor")
    col_ubic = col_idx("ubica")
    col_desc = col_idx("descrip")
    col_estado = col_idx("estado")
    col_cant = col_idx("cantid")
    col_marca = col_idx("marca")
    col_modelo = col_idx("modelo")
    col_serie = col_idx("serie")
    col_notas = col_idx("nota")

    if col_nombre is None:
        print("ERROR: No se encontró la columna 'nombre' en el Excel.")
        return

    # Cache de categorías
    cat_cache = {}
    for cat in Categoria.query.all():
        cat_cache[cat.nombre.lower()] = cat.id

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        nombre = vals[col_nombre] if col_nombre is not None else None
        if not nombre:
            continue

        # Resolver categoría
        cat_nombre = str(vals[col_cat]).strip() if col_cat is not None and vals[col_cat] else "Otros"
        cat_id = cat_cache.get(cat_nombre.lower())
        if not cat_id:
            nueva_cat = Categoria(nombre=cat_nombre, es_global=False)
            nueva_cat.barrios = [barrio]
            db.session.add(nueva_cat)
            db.session.flush()
            cat_cache[cat_nombre.lower()] = nueva_cat.id
            cat_id = nueva_cat.id

        item = Item(
            nombre=str(nombre).strip(),
            codigo=str(vals[col_codigo]).strip() if col_codigo is not None and vals[col_codigo] else None,
            descripcion=str(vals[col_desc]).strip() if col_desc and vals[col_desc] else "",
            categoria_id=cat_id,
            barrio_id=barrio_id,
            ubicacion=str(vals[col_ubic]).strip() if col_ubic and vals[col_ubic] else "",
            estado=str(vals[col_estado]).strip() if col_estado and vals[col_estado] else "Operativo",
            cantidad=int(vals[col_cant]) if col_cant and vals[col_cant] else 1,
            marca=str(vals[col_marca]).strip() if col_marca and vals[col_marca] else "",
            modelo=str(vals[col_modelo]).strip() if col_modelo and vals[col_modelo] else "",
            numero_serie=str(vals[col_serie]).strip() if col_serie and vals[col_serie] else "",
            notas=str(vals[col_notas]).strip() if col_notas and vals[col_notas] else "",
            created_by=user_id,
        )
        db.session.add(item)
        db.session.flush()

        db.session.add(Historial(
            item_id=item.id,
            user_id=user_id,
            accion="alta",
            detalle="Importado desde Excel",
        ))
        count += 1

    db.session.commit()
    print(f"OK: {count} items importados al barrio {barrio_id} ({barrio.nombre})")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python scripts/import_excel.py <archivo.xlsx> <barrio_id>")
        sys.exit(1)

    with app.app_context():
        importar(sys.argv[1], int(sys.argv[2]))
